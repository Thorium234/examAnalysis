# exams/utils/spreadsheet.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from students.models import Student, StudentSubjectEnrollment
from exams.models import Exam, PaperResult, SubjectPaper

class SpreadsheetTemplate:
    def __init__(self, exam, subject, papers=None):
        self.exam = exam
        self.subject = subject
        self.papers = papers if isinstance(papers, list) else ([papers] if papers else [])
        self.wb = Workbook()
        self.summary_sheet = self.wb.active
        self.summary_sheet.title = 'Summary'
        self.validation_errors = []
        self.processing_progress = {
            'total': 0,
            'processed': 0,
            'success': 0,
            'errors': 0
        }
        
        # Define styles
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.locked_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_template(self):
        """Generate a multi-sheet workbook with papers and summary"""
        # Get eligible students
        self.students = Student.objects.filter(
            form_level__in=self.exam.participating_forms.all(),
            studentsubjectenrollment__subject=self.subject,
            studentsubjectenrollment__is_enrolled=True
        ).exclude(
            paperresult__exam=self.exam,
            paperresult__subject=self.subject
        ).order_by('admission_number')

        # Create paper sheets
        if self.papers:
            for paper in self.papers:
                self._create_paper_sheet(paper)
        else:
            # Create single sheet for direct entry
            self._create_paper_sheet(None)
            
        # Create summary sheet
        self._create_summary_sheet()
        
        return self.wb
        
    def _create_paper_sheet(self, paper):
        """Create a worksheet for a specific paper or direct entry"""
        sheet_name = paper.name if paper else "Direct Entry"
        ws = self.wb.create_sheet(sheet_name)
        
        # Set up protection
        ws.protection.sheet = True
        ws.protection.password = 'examsheet'
        
        # Add metadata
        ws.row_dimensions[1].hidden = True
        self._add_metadata(ws, paper)
        
        # Add headers
        headers = ['Admission Number', 'Student Name', 'Stream']
        if paper:
            headers.append(f'Marks (out of {paper.max_marks})')
        else:
            headers.append('Marks (out of 100)')
        headers.append('Status (P/A/D)')
        headers.extend(['Entry Date', 'Comments'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
            
        # Add student data
        for row, student in enumerate(self.students, 3):
            self._add_student_row(ws, row, student, paper)
            
        # Set column widths
        column_widths = {
            'A': 15,  # Admission Number
            'B': 30,  # Name
            'C': 10,  # Stream
            'D': 15,  # Marks
            'E': 10,  # Status
            'F': 15,  # Entry Date
            'G': 30,  # Comments
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
    def _create_summary_sheet(self):
        """Create a summary sheet with statistics and validation status"""
        ws = self.summary_sheet
        
        # Add exam and subject info
        ws['A1'] = 'Exam:'
        ws['B1'] = self.exam.name
        ws['A2'] = 'Subject:'
        ws['B2'] = self.subject.name
        
        # Add paper summary
        ws['A4'] = 'Papers Summary:'
        headers = ['Paper', 'Max Marks', 'Students', 'Entries', 'Missing', 'Invalid']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = Font(bold=True)
            
        row = 6
        for paper in self.papers:
            ws.cell(row=row, column=1, value=paper.name)
            ws.cell(row=row, column=2, value=paper.max_marks)
            ws.cell(row=row, column=3, value=len(self.students))
            row += 1
            
        # Add validation section
        ws['A10'] = 'Validation Rules:'
        rules = [
            'All student information must remain unchanged',
            'Marks must be within allowed range for each paper',
            'Status must be P (Present), A (Absent), or D (Disqualified)',
            'Present students must have valid marks',
            'Absent/Disqualified students should not have marks'
        ]
        
        for i, rule in enumerate(rules, 11):
            ws[f'A{i}'] = f'• {rule}'
        
        # Add student data
        for row, student in enumerate(students, 3):
            # Locked cells
            self.ws.cell(row=row, column=1, value=student.admission_number).protection = Protection(locked=True)
            self.ws.cell(row=row, column=2, value=student.name).protection = Protection(locked=True)
            self.ws.cell(row=row, column=3, value=student.stream).protection = Protection(locked=True)
            
            # Mark entry cell
            marks_cell = self.ws.cell(row=row, column=4)
            marks_cell.protection = Protection(locked=False)
            
            # Status cell with data validation
            status_cell = self.ws.cell(row=row, column=5)
            status_cell.protection = Protection(locked=False)
            
            # Apply styles
            for col in range(1, 6):
                cell = self.ws.cell(row=row, column=col)
                cell.border = self.border
                if col < 4:  # Locked cells
                    cell.fill = self.locked_fill
        
        # Add data validation for status column
        status_range = f'E3:E{len(students) + 2}'
        self.ws.data_validation = {
            status_range: 'P,A,D',
            'type': 'list',
            'allowBlank': False,
            'showInputMessage': True,
            'promptTitle': 'Status',
            'prompt': 'P=Present, A=Absent, D=Disqualified'
        }
        
        # Set column widths
        self.ws.column_dimensions['A'].width = 15  # Admission Number
        self.ws.column_dimensions['B'].width = 30  # Name
        self.ws.column_dimensions['C'].width = 10  # Stream
        self.ws.column_dimensions['D'].width = 15  # Marks
        self.ws.column_dimensions['E'].width = 10  # Status
        
        return self.wb
    
    def _add_metadata(self):
        """Add hidden metadata for validation during upload"""
        metadata = {
            'exam_id': self.exam.id,
            'subject_id': self.subject.id,
            'paper_id': self.paper.id if self.paper else '',
            'max_marks': self.paper.max_marks if self.paper else 100,
            'template_version': '1.0'
        }
        
        for col, (key, value) in enumerate(metadata.items(), 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = f"{key}:{value}"
    
    def validate_spreadsheet(self, file):
        """Validate uploaded spreadsheet against template format with detailed checking"""
        self.validation_errors = []
        try:
            # Read all sheets
            xlsx = pd.ExcelFile(file)
            sheets_data = {}
            metadata = {}
            
            for sheet_name in xlsx.sheet_names:
                if sheet_name == 'Summary':
                    continue
                    
                df = pd.read_excel(xlsx, sheet_name=sheet_name, header=1)
                sheets_data[sheet_name] = df
                
                # Extract metadata from hidden row
                meta_row = pd.read_excel(xlsx, sheet_name=sheet_name, nrows=1)
                metadata[sheet_name] = self._parse_metadata(meta_row)
                
            # Validate each sheet
            for sheet_name, df in sheets_data.items():
                self._validate_sheet(df, metadata[sheet_name], sheet_name)
                
            if self.validation_errors:
                return False, self._format_validation_errors()
                
            return True, "Validation successful"
            
        except Exception as e:
            return False, f"Error validating spreadsheet: {str(e)}"
            
    def _validate_sheet(self, df, metadata, sheet_name):
        """Perform detailed validation on a single sheet"""
        # Required columns check
        required_columns = [
            'Admission Number', 'Student Name', 'Stream',
            'Marks', 'Status', 'Entry Date', 'Comments'
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            self.validation_errors.append(
                f"Sheet '{sheet_name}' is missing columns: {', '.join(missing_columns)}"
            )
            
        # Student data integrity
        for idx, row in df.iterrows():
            self._validate_student_row(row, metadata, sheet_name, idx + 3)
            
        # Check for duplicate entries
        duplicates = df['Admission Number'].duplicated()
        if duplicates.any():
            duplicate_numbers = df[duplicates]['Admission Number'].tolist()
            self.validation_errors.append(
                f"Sheet '{sheet_name}' has duplicate entries for: {', '.join(duplicate_numbers)}"
            )
            
    def _validate_student_row(self, row, metadata, sheet_name, row_num):
        """Validate a single student row with detailed checks"""
        # Status validation
        if pd.notna(row['Status']) and row['Status'] not in ['P', 'A', 'D']:
            self.validation_errors.append(
                f"Invalid status '{row['Status']}' in {sheet_name} row {row_num}"
            )
            
        # Marks validation
        if row['Status'] == 'P':
            if pd.isna(row['Marks']):
                self.validation_errors.append(
                    f"Missing marks for present student in {sheet_name} row {row_num}"
                )
            elif not (isinstance(row['Marks'], (int, float)) and 
                     0 <= row['Marks'] <= float(metadata.get('max_marks', 100))):
                self.validation_errors.append(
                    f"Invalid marks '{row['Marks']}' in {sheet_name} row {row_num}"
                )
        elif pd.notna(row['Marks']):
            self.validation_errors.append(
                f"Marks should not be entered for {row['Status']} status in {sheet_name} row {row_num}"
            )
            
        # Student existence check
        try:
            student = Student.objects.get(admission_number=row['Admission Number'])
            if student.name != row['Student Name'] or student.stream != row['Stream']:
                self.validation_errors.append(
                    f"Student information mismatch in {sheet_name} row {row_num}"
                )
        except Student.DoesNotExist:
            self.validation_errors.append(
                f"Invalid admission number '{row['Admission Number']}' in {sheet_name} row {row_num}"
            )
    
    def process_spreadsheet(self, file):
        """Process uploaded spreadsheet with progress tracking and summary generation"""
        try:
            xlsx = pd.ExcelFile(file)
            self.processing_progress['total'] = sum(
                len(pd.read_excel(xlsx, sheet_name=name, header=1))
                for name in xlsx.sheet_names if name != 'Summary'
            )
            
            all_results = []
            sheet_summaries = {}
            
            for sheet_name in xlsx.sheet_names:
                if sheet_name == 'Summary':
                    continue
                    
                df = pd.read_excel(xlsx, sheet_name=sheet_name, header=1)
                meta_row = pd.read_excel(xlsx, sheet_name=sheet_name, nrows=1)
                metadata = self._parse_metadata(meta_row)
                
                results, summary = self._process_sheet(df, metadata, sheet_name)
                all_results.extend(results)
                sheet_summaries[sheet_name] = summary
                
            # Bulk create all results
            PaperResult.objects.bulk_create(all_results)
            
            # Generate final summary
            return self._generate_processing_summary(sheet_summaries)
            
        except Exception as e:
            raise Exception(f"Error processing spreadsheet: {str(e)}")
    
    def _process_sheet(self, df, metadata, sheet_name):
        """Process a single sheet and generate summary"""
        results = []
        summary = {
            'total': len(df),
            'processed': 0,
            'success': 0,
            'errors': 0,
            'absent': 0,
            'disqualified': 0,
            'marks_distribution': {
                'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0
            }
        }
        
        for idx, row in df.iterrows():
            try:
                if pd.isna(row['Status']):
                    continue
                    
                student = Student.objects.get(admission_number=row['Admission Number'])
                result = PaperResult(
                    exam_id=metadata['exam_id'],
                    student=student,
                    subject_id=metadata['subject_id'],
                    paper_id=metadata.get('paper_id'),
                    marks=row['Marks'] if row['Status'] == 'P' else None,
                    status=row['Status'],
                    comments=row.get('Comments', '')
                )
                
                # Update summary statistics
                if row['Status'] == 'P':
                    grade = self._calculate_grade(row['Marks'])
                    summary['marks_distribution'][grade[0]] += 1
                    summary['success'] += 1
                elif row['Status'] == 'A':
                    summary['absent'] += 1
                elif row['Status'] == 'D':
                    summary['disqualified'] += 1
                    
                results.append(result)
                summary['processed'] += 1
                
            except Exception as e:
                summary['errors'] += 1
                self.validation_errors.append(
                    f"Error in {sheet_name} row {idx + 3}: {str(e)}"
                )
                
            # Update progress
            self.processing_progress['processed'] += 1
            
        return results, summary
    
    def _generate_processing_summary(self, sheet_summaries):
        """Generate a detailed processing summary"""
        total_summary = {
            'total_entries': sum(s['total'] for s in sheet_summaries.values()),
            'successful': sum(s['success'] for s in sheet_summaries.values()),
            'errors': sum(s['errors'] for s in sheet_summaries.values()),
            'absent': sum(s['absent'] for s in sheet_summaries.values()),
            'disqualified': sum(s['disqualified'] for s in sheet_summaries.values()),
            'marks_distribution': {},
            'sheet_details': sheet_summaries
        }
        
        # Aggregate marks distribution
        for grade in ['A', 'B', 'C', 'D', 'E']:
            total_summary['marks_distribution'][grade] = sum(
                s['marks_distribution'][grade] for s in sheet_summaries.values()
            )
            
        return total_summary
        
    def _calculate_grade(self, marks):
        """Calculate grade for summary statistics"""
        if marks >= 80:
            return 'A'
        elif marks >= 65:
            return 'B'
        elif marks >= 50:
            return 'C'
        elif marks >= 40:
            return 'D'
        else:
            return 'E'